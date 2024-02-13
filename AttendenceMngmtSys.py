import tkinter as tk
import tkinter.ttk as ttk
from tkinter import *
from PIL import ImageTk, Image
from tkinter import messagebox
import pandas as pd
import openpyxl

window = tk.Tk()
window.title("Attendance Management System")
window.geometry('550x500')

def login():
    username = entry_username.get()
    password = entry_password.get()
    if username == "admin" and password == "poornima":
        messagebox.showinfo("Success", "Login successful")
        window.deiconify() 
        login_window.destroy()
    elif username == "faculty" and password == "facultypu":
        messagebox.showinfo("Success", "Login successful, welcome faculty")
        display_data()
        login_window.destroy()
    else:
        messagebox.showinfo("Failed!", "Enter correct Username of Password")


login_window = tk.Toplevel(window)
login_window.title("Login")
login_window.geometry('550x500')
login_window.protocol("WM_DELETE_WINDOW", login_window.quit) 


login_window.resizable(False, False)


canvas = tk.Canvas(login_window, width=550, height=500)
canvas.pack(fill='both', expand=True)

image = Image.open("background.jpg")
background_image = ImageTk.PhotoImage(image)
canvas.create_image(0, 0, image=background_image, anchor='nw')

frame = tk.Frame(canvas, bg='white')
frame.place(relx=0.5, rely=0.5, anchor='center')

label_username = tk.Label(frame, text="Username", bg="purple", fg="white")
label_username.grid(row=0, column=0)
entry_username = tk.Entry(frame)
entry_username.grid(row=0, column=1)

label_password = tk.Label(frame, text="Password", bg="blue", fg="white")
label_password.grid(row=1, column=0, padx=5, pady=10)
entry_password = tk.Entry(frame, show="*")
entry_password.grid(row=1, column=1, padx=5, pady=10)

button_login = tk.Button(frame, text="Login", command=login, bg="green", fg="white")
button_login.grid(row=2, column=0, columnspan=2, pady=10)

window.withdraw()

student_data = {}

def add_details():
    name = entry_name.get()
    reg_no = entry_reg_no.get()
    section = entry_section.get()
    subjects = ",".join(entry_subjects.get().split(","))
    student_data[reg_no] = [name, section, subjects]
    entry_name.delete(0, END)
    entry_reg_no.delete(0, END)
    entry_section.delete(0, END)
    entry_subjects.delete(0, END)
    messagebox.showinfo("Success", "Student details added successfully")

def display_data():
    window3 = tk.Toplevel(window)
    window3.title("Display Student Data")
    window3.geometry('900x500')
    window3.resizable(False, False)
    tree = ttk.Treeview(window3, columns=("column1", "column2", "column3", "column4"), show='headings')
    tree.heading("#1", text="Name")
    tree.heading("#2", text="Registration Number")
    tree.heading("#3", text="Section")
    tree.heading("#4", text="Subjects")
    tree.grid(row=0, column=0)

    try:
        df = pd.read_excel("StudentData.xlsx", index_col='RegNo')
        for index, row in df.iterrows():
            name = row['Name']
            reg_no = index
            section = row['Section']
            subjects = row['Subjects']
            tree.insert("", END, values=(name, reg_no, section, subjects))
    except FileNotFoundError:
        messagebox.showinfo("", "No File found Creating a new File")

    for key, val in student_data.items():

        name = val[0]
        reg_no = key
        section = val[1]
        subjects = ",".join(val[2])
        tree.insert("", END, values=(name, reg_no, section, subjects))

    def mark_attendance():
        items = tree.selection()
        for item in items:
            item_data = tree.item(item)["values"]
            reg_no = item_data[1]
            if var.get() == 1:
                present = +1  
            else:
                present = 0 
            if reg_no in student_data:
                if len(student_data[reg_no]) == 3:
                    student_data[reg_no].append([present])
                else:
                    student_data[reg_no][3].append(present)
        messagebox.showinfo("Success","Attendance marked successfully")
    button_mark = Button(window3,text="Mark Attendance",command=mark_attendance)
    button_mark.grid(column=0,row=1)
    var = IntVar()
    radio_present = Radiobutton(window3, text="Present", variable=var, value=1).grid(column=0,row=2)
    radio_absent = Radiobutton(window3, text="Absent", variable=var, value=2).grid(column=0,row=3)

    
    def save_student_data():
        save_data()
        messagebox.showinfo("Success", "Student data saved successfully.")

    button_mark = Button(window3, text="Mark Attendance", command=mark_attendance)
    button_mark.grid(column=0, row=1)
    var = IntVar()
    radio_present = Radiobutton(window3, text="Present", variable=var, value=1).grid(column=0, row=2)
    radio_absent = Radiobutton(window3, text="Absent", variable=var, value=2).grid(column=0, row=3)
    
    button_save = Button(window3, text="Save Data", command=save_student_data)
    button_save.grid(column=0, row=4)

def save_data():
    try:
        wb = openpyxl.load_workbook("StudentData.xlsx")
        ws = wb.active
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Sheet1'

    headers = ['Name', 'RegNo', 'Section', 'Subjects', 'Attendance']
    for col_num, header in enumerate(headers, 1):
        if ws.cell(row=1, column=col_num).value != header:
            ws.cell(row=1, column=col_num).value = header

    next_row = ws.max_row + 1
    for reg_no, data in student_data.items():
        if ws.cell(row=next_row, column=2).value == reg_no:
            continue  
        name, section, subjects = data[:3]
        attendance = ','.join(map(str, data[3]))
        ws.cell(row=next_row, column=1).value = name
        ws.cell(row=next_row, column=2).value = reg_no
        ws.cell(row=next_row, column=3).value = section
        ws.cell(row=next_row, column=4).value = subjects
        ws.cell(row=next_row, column=5).value = attendance
        next_row += 1

    wb.save("StudentData.xlsx")
    messagebox.showinfo("Success","Data saved successfully to existing file")
        
bg_image = PhotoImage(file="background.jpg")

canvas = Canvas(window, width=bg_image.width(), height=bg_image.height())
canvas.grid(row=0, column=0)

canvas.create_image(0, 0, image=bg_image, anchor="nw")

entry_frame = Frame(canvas, bg="white")
entry_frame.place(relx=0.5, rely=0.5, anchor="center")

label_name = Label(entry_frame, text="Name", font=("Arial", 12), bg="blue", fg="white")
label_name.grid(column=0, row=0, sticky="w")
entry_name = Entry(entry_frame, width=30, bg="white", fg="black")
entry_name.grid(column=1, row=0, padx=5, pady=10)

label_reg_no = Label(entry_frame, text="Registration Number", font=("Arial", 12), bg="purple", fg="white")
label_reg_no.grid(column=0, row=1, sticky="w")
entry_reg_no = Entry(entry_frame, width=30, bg="white", fg="black")
entry_reg_no.grid(column=1, row=1, padx=5, pady=10)

label_section = Label(entry_frame, text="Section", font=("Arial", 12), bg="orange", fg="white")
label_section.grid(column=0, row=2, sticky="w")
entry_section = Entry(entry_frame, width=30, bg="white", fg="black")
entry_section.grid(column=1, row=2, padx=5, pady=10)

label_subjects = Label(entry_frame, text="Subjects", font=("Arial", 12), bg="green", fg="white")
label_subjects.grid(column=0, row=3, sticky="w")
entry_subjects = Entry(entry_frame, width=30, bg="white", fg="black")
entry_subjects.grid(column=1, row=3, padx=5, pady=10)

button_add = Button(entry_frame, text="Add Student", command=add_details, bg="green", fg="white")
button_add.grid(column=1, row=4, pady=10)

button_display = Button(entry_frame, text="Display Data", command=display_data, bg="blue", fg="white")
button_display.grid(column=1, row=5, pady=10)

button_save = Button(entry_frame, text="Save Data", command=save_data, bg="purple", fg="white")
button_save.grid(column=1, row=6, pady=10)

window.geometry("{}x{}".format(bg_image.width(), bg_image.height()))

window.resizable(False, False)


window.mainloop()
